from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, func, and_
from uuid import uuid4
from datetime import datetime
from typing import List, Optional
from decimal import Decimal
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.application.dtos import FundoResponseDTO, FundoUpdateDTO
from app.infrastructure.database import get_db
from app.infrastructure.database.models import FundoCarregamentoModel, FundoModel, UserModel, MovimentoFinanceiroModel
from app.infrastructure.repositories import FundoRepository
from app.infrastructure.export import aplicar_cabecalho_empresa
from app.infrastructure.auth.dependencies import get_current_user, require_financeiro
from app.domain.entities import User, Fundo

router = APIRouter()


async def _get_or_create(repo: FundoRepository, company_id, tipo: str, db) -> Fundo:
    fundo = await repo.get_by_company_and_tipo(company_id, tipo)
    if not fundo:
        entity = Fundo(
            id=uuid4(),
            company_id=company_id,
            tipo=tipo,
            valor_disponivel=Decimal("0.00"),
            acumulado=Decimal("0.00"),
            saldo_atual=Decimal("0.00"),
        )
        fundo = await repo.create(entity)
        await db.commit()
    return fundo


async def _fund_totals(db: AsyncSession, company_id, fundo_tipo: str):
    try:
        r_e = await db.execute(
            select(func.coalesce(func.sum(MovimentoFinanceiroModel.valor), 0)).where(and_(
                MovimentoFinanceiroModel.company_id == company_id,
                MovimentoFinanceiroModel.fundo_tipo == fundo_tipo,
                MovimentoFinanceiroModel.tipo_movimento == "entrada",
                MovimentoFinanceiroModel.deleted_at == None,
            ))
        )
        r_s = await db.execute(
            select(func.coalesce(func.sum(MovimentoFinanceiroModel.valor), 0)).where(and_(
                MovimentoFinanceiroModel.company_id == company_id,
                MovimentoFinanceiroModel.fundo_tipo == fundo_tipo,
                MovimentoFinanceiroModel.tipo_movimento == "saida",
                MovimentoFinanceiroModel.deleted_at == None,
            ))
        )
        return Decimal(str(r_e.scalar_one() or 0)), Decimal(str(r_s.scalar_one() or 0))
    except Exception:
        try:
            await db.rollback()
        except Exception:
            pass
        return Decimal("0"), Decimal("0")


@router.get("")
async def obter_fundos(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    repo = FundoRepository(db)
    bcs = await _get_or_create(repo, current_user.company_id, "BCS", db)
    bfa = await _get_or_create(repo, current_user.company_id, "BFA", db)
    bcs_e, bcs_s = await _fund_totals(db, current_user.company_id, "BCS")
    bfa_e, bfa_s = await _fund_totals(db, current_user.company_id, "BFA")
    return {"bcs": _to_dto(bcs, bcs_e, bcs_s), "bfa": _to_dto(bfa, bfa_e, bfa_s)}


@router.put("", response_model=FundoResponseDTO)
async def atualizar_fundo(
    body: FundoUpdateDTO,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_financeiro),
):
    repo = FundoRepository(db)
    fundo = await _get_or_create(repo, current_user.company_id, body.tipo, db)

    valor_anterior = Decimal(str(fundo.valor_disponivel or 0))
    novo_disponivel = Decimal(str(body.valor_disponivel))
    novo_saldo = novo_disponivel - Decimal(str(fundo.acumulado or 0))

    updated_entity = Fundo(
        id=fundo.id,
        company_id=fundo.company_id,
        tipo=fundo.tipo,
        data=fundo.data,
        descricao=fundo.descricao,
        valor_disponivel=novo_disponivel,
        acumulado=Decimal(str(fundo.acumulado or 0)),
        saldo_atual=novo_saldo,
        observacao=body.observacao or fundo.observacao,
    )
    updated = await repo.update(fundo.id, updated_entity)

    carregamento = FundoCarregamentoModel(
        id=uuid4(),
        company_id=current_user.company_id,
        fundo_id=fundo.id,
        user_id=current_user.id,
        valor_anterior=valor_anterior,
        valor_novo=novo_disponivel,
        observacao=body.observacao,
        origem=body.origem,
        created_at=datetime.utcnow(),
    )
    db.add(carregamento)
    await db.commit()
    return _to_dto(updated)


@router.get("/export")
async def exportar_historico_excel(
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    repo = FundoRepository(db)
    bcs = await repo.get_by_company_and_tipo(current_user.company_id, "BCS")
    bfa = await repo.get_by_company_and_tipo(current_user.company_id, "BFA")

    q = (
        select(FundoCarregamentoModel, UserModel.full_name, FundoModel.tipo.label("fundo_tipo"))
        .join(UserModel, FundoCarregamentoModel.user_id == UserModel.id)
        .outerjoin(FundoModel, FundoCarregamentoModel.fundo_id == FundoModel.id)
        .where(FundoCarregamentoModel.company_id == current_user.company_id)
    )
    if data_inicio:
        q = q.where(FundoCarregamentoModel.created_at >= data_inicio)
    if data_fim:
        q = q.where(FundoCarregamentoModel.created_at <= data_fim)
    q = q.order_by(desc(FundoCarregamentoModel.created_at))
    result = await db.execute(q)
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico de Fundos"

    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    summary_fill = PatternFill("solid", fgColor="EFF6FF")
    summary_font = Font(bold=True, size=11)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    # Cabeçalho da empresa (logo + nome + NIF/morada + título)
    periodo_txt = ""
    if data_inicio: periodo_txt += f"De {data_inicio.strftime('%d/%m/%Y')} "
    if data_fim: periodo_txt += f"a {data_fim.strftime('%d/%m/%Y')}"
    header_row = await aplicar_cabecalho_empresa(
        ws, db, current_user.company_id,
        titulo="Histórico de Carregamentos de Fundos",
        subtitulo=periodo_txt or None,
        n_colunas=8,
    )

    headers = ["Nº", "Data / Hora", "Utilizador", "Fundo", "Valor Anterior (AOA)", "Valor Novo (AOA)", "Variação (AOA)", "Observação"]
    col_widths = [5, 20, 28, 8, 22, 22, 22, 40]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[header_row].height = 20

    for idx, row in enumerate(rows, header_row + 1):
        c = row.FundoCarregamentoModel
        v_ant = float(c.valor_anterior or 0)
        v_novo = float(c.valor_novo)
        diff = v_novo - v_ant
        row_data = [idx - header_row, c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else "", row.full_name, row.fundo_tipo or "BCS", v_ant, v_novo, diff, c.observacao or ""]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin
            if col == 4:
                cell.alignment = center
            if col in (5, 6, 7):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            if col == 7:
                cell.font = Font(color="166534", bold=True) if diff >= 0 else Font(color="991B1B", bold=True)

    blank_row = header_row + len(rows) + 1
    ws.row_dimensions[blank_row].height = 8
    summary_start = blank_row + 1

    for i, (label, value) in enumerate([
        ("Total de Carregamentos", len(rows)),
        ("", ""),
        ("BCS — Valor Disponível (AOA)", float(bcs.valor_disponivel) if bcs else 0.0),
        ("BCS — Acumulado (AOA)", float(bcs.acumulado) if bcs else 0.0),
        ("BCS — Saldo Actual (AOA)", float(bcs.saldo_atual) if bcs else 0.0),
        ("", ""),
        ("BFA — Valor Disponível (AOA)", float(bfa.valor_disponivel) if bfa else 0.0),
        ("BFA — Acumulado (AOA)", float(bfa.acumulado) if bfa else 0.0),
        ("BFA — Saldo Actual (AOA)", float(bfa.saldo_atual) if bfa else 0.0),
    ]):
        r = summary_start + i
        lc = ws.cell(row=r, column=7, value=label)
        lc.font = summary_font; lc.fill = summary_fill; lc.border = thin
        lc.alignment = Alignment(horizontal="right")
        vc = ws.cell(row=r, column=8, value=value)
        vc.font = summary_font; vc.fill = summary_fill; vc.border = thin
        if isinstance(value, float):
            vc.number_format = '#,##0.00'
        vc.alignment = Alignment(horizontal="right")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"historico_fundos_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/historico")
async def historico_fundo(
    limit: int = 20,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[dict]:
    result = await db.execute(
        select(FundoCarregamentoModel, UserModel.full_name, FundoModel.tipo.label("fundo_tipo"))
        .join(UserModel, FundoCarregamentoModel.user_id == UserModel.id)
        .outerjoin(FundoModel, FundoCarregamentoModel.fundo_id == FundoModel.id)
        .where(FundoCarregamentoModel.company_id == current_user.company_id)
        .order_by(desc(FundoCarregamentoModel.created_at))
        .limit(limit)
    )
    rows = result.all()
    return [
        {
            "id": str(row.FundoCarregamentoModel.id),
            "valor_anterior": float(row.FundoCarregamentoModel.valor_anterior or 0),
            "valor_novo": float(row.FundoCarregamentoModel.valor_novo),
            "observacao": row.FundoCarregamentoModel.observacao,
            "origem": row.FundoCarregamentoModel.origem,
            "user_name": row.full_name,
            "fundo_tipo": row.fundo_tipo or "BCS",
            "created_at": row.FundoCarregamentoModel.created_at.isoformat(),
        }
        for row in rows
    ]


def _to_dto(f: Fundo, total_entradas: Decimal = Decimal("0"), total_saidas: Decimal = Decimal("0")) -> FundoResponseDTO:
    return FundoResponseDTO(
        id=f.id,
        company_id=f.company_id,
        tipo=f.tipo,
        data=f.data,
        descricao=f.descricao,
        valor_disponivel=Decimal(str(f.valor_disponivel or 0)),
        acumulado=Decimal(str(f.acumulado or 0)),
        saldo_atual=Decimal(str(f.saldo_atual or 0)),
        total_entradas=total_entradas,
        total_saidas=total_saidas,
        observacao=f.observacao,
        created_at=f.created_at,
        updated_at=f.updated_at,
    )


__all__ = ["router"]
