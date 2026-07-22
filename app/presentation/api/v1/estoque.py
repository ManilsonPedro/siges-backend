"""Estoque: saldos, movimentos, kardex, alertas.

Toda alteração de saldo passa pelo :mod:`app.domain.services.stock_service`.
"""
from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from typing import List, Optional
from uuid import UUID

import io

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import select, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.domain.entities import User
from app.domain.services import stock_service
from app.infrastructure.audit import write_audit
from app.infrastructure.auth.dependencies import get_current_user
from app.infrastructure.database import get_db
from app.infrastructure.database.models import (
    ArmazemModel,
    ProdutoModel,
    StockMovimentoModel,
    StockSaldoModel,
)
from app.infrastructure.export import aplicar_cabecalho_empresa


router = APIRouter()


# ─── DTOs ────────────────────────────────────────────────────────────


class SaldoResponseDTO(BaseModel):
    produto_id: UUID
    produto_sku: str
    produto_nome: str
    armazem_id: UUID
    armazem_codigo: str
    armazem_nome: str
    qtd_actual: Decimal
    qtd_reservada: Decimal
    qtd_disponivel: Decimal
    stock_minimo: Decimal
    abaixo_minimo: bool


class MovimentoResponseDTO(BaseModel):
    id: UUID
    produto_id: UUID
    armazem_origem_id: Optional[UUID] = None
    armazem_destino_id: Optional[UUID] = None
    tipo: str
    quantidade: Decimal
    custo_unitario: Optional[Decimal] = None
    documento_ref_tipo: Optional[str] = None
    documento_ref_id: Optional[str] = None
    motivo: Optional[str] = None
    estornado_de: Optional[UUID] = None
    created_by: Optional[UUID] = None
    created_at: datetime

    class Config:
        from_attributes = True


class EntradaDTO(BaseModel):
    produto_id: UUID
    armazem_id: UUID
    quantidade: Decimal = Field(..., gt=0)
    custo_unitario: Optional[Decimal] = Field(None, ge=0)
    tipo: str = Field(default="entrada_compra",
                      pattern="^(entrada_compra|entrada_producao|entrada_ajuste)$")
    motivo: Optional[str] = None


class SaidaDTO(BaseModel):
    produto_id: UUID
    armazem_id: UUID
    quantidade: Decimal = Field(..., gt=0)
    tipo: str = Field(default="saida_perda",
                      pattern="^(saida_perda|saida_ajuste)$")
    motivo: Optional[str] = None


class TransferenciaDTO(BaseModel):
    produto_id: UUID
    armazem_origem_id: UUID
    armazem_destino_id: UUID
    quantidade: Decimal = Field(..., gt=0)
    motivo: Optional[str] = None


class StockMinimoDTO(BaseModel):
    produto_id: UUID
    armazem_id: UUID
    stock_minimo: Decimal = Field(..., ge=0)


# ─── Saldos ──────────────────────────────────────────────────────────


@router.get("/saldos", response_model=List[SaldoResponseDTO])
async def list_saldos(
    armazem_id: Optional[UUID] = None,
    produto_id: Optional[UUID] = None,
    abaixo_minimo: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    stmt = (
        select(StockSaldoModel, ProdutoModel, ArmazemModel)
        .join(ProdutoModel, ProdutoModel.id == StockSaldoModel.produto_id)
        .join(ArmazemModel, ArmazemModel.id == StockSaldoModel.armazem_id)
        .where(StockSaldoModel.company_id == current_user.company_id)
        .where(ProdutoModel.deleted_at.is_(None))
        .where(ArmazemModel.deleted_at.is_(None))
    )
    if armazem_id:
        stmt = stmt.where(StockSaldoModel.armazem_id == armazem_id)
    if produto_id:
        stmt = stmt.where(StockSaldoModel.produto_id == produto_id)
    stmt = stmt.order_by(ProdutoModel.nome, ArmazemModel.codigo)
    r = await db.execute(stmt)
    out: List[SaldoResponseDTO] = []
    for saldo, prod, arm in r.all():
        disp = Decimal(saldo.qtd_actual) - Decimal(saldo.qtd_reservada)
        ab_min = Decimal(saldo.qtd_actual) < Decimal(saldo.stock_minimo) \
            and Decimal(saldo.stock_minimo) > 0
        if abaixo_minimo and not ab_min:
            continue
        out.append(SaldoResponseDTO(
            produto_id=prod.id, produto_sku=prod.sku, produto_nome=prod.nome,
            armazem_id=arm.id, armazem_codigo=arm.codigo, armazem_nome=arm.nome,
            qtd_actual=Decimal(saldo.qtd_actual),
            qtd_reservada=Decimal(saldo.qtd_reservada),
            qtd_disponivel=disp,
            stock_minimo=Decimal(saldo.stock_minimo),
            abaixo_minimo=ab_min,
        ))
    return out


@router.post("/saldos/stock-minimo", status_code=204)
async def set_stock_minimo(
    body: StockMinimoDTO,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    r = await db.execute(
        select(StockSaldoModel)
        .where(StockSaldoModel.produto_id == body.produto_id)
        .where(StockSaldoModel.armazem_id == body.armazem_id)
    )
    s = r.scalar_one_or_none()
    if not s:
        # cria saldo zerado para guardar o mínimo
        s = await stock_service._get_or_create_saldo(
            db, company_id=current_user.company_id,
            produto_id=body.produto_id, armazem_id=body.armazem_id,
        )
    s.stock_minimo = body.stock_minimo
    await db.commit()


@router.get("/alertas/stock-minimo", response_model=List[SaldoResponseDTO])
async def alertas_stock_minimo(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return await list_saldos(
        armazem_id=None, produto_id=None, abaixo_minimo=True,
        db=db, current_user=current_user,
    )


# ─── Movimentos ──────────────────────────────────────────────────────


@router.get("/movimentos", response_model=List[MovimentoResponseDTO])
async def list_movimentos(
    data_de: Optional[datetime] = None,
    data_ate: Optional[datetime] = None,
    tipo: Optional[str] = None,
    produto_id: Optional[UUID] = None,
    armazem_id: Optional[UUID] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    stmt = (
        select(StockMovimentoModel)
        .where(StockMovimentoModel.company_id == current_user.company_id)
    )
    if data_de:
        stmt = stmt.where(StockMovimentoModel.created_at >= data_de)
    if data_ate:
        stmt = stmt.where(StockMovimentoModel.created_at <= data_ate)
    if tipo:
        stmt = stmt.where(StockMovimentoModel.tipo == tipo)
    if produto_id:
        stmt = stmt.where(StockMovimentoModel.produto_id == produto_id)
    if armazem_id:
        stmt = stmt.where(or_(
            StockMovimentoModel.armazem_origem_id == armazem_id,
            StockMovimentoModel.armazem_destino_id == armazem_id,
        ))
    stmt = stmt.order_by(StockMovimentoModel.created_at.desc()) \
               .offset((page - 1) * page_size).limit(page_size)
    r = await db.execute(stmt)
    return list(r.scalars().all())


@router.post("/movimentos/entrada", response_model=MovimentoResponseDTO, status_code=201)
async def movimento_entrada(
    req: Request,
    body: EntradaDTO,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mov = await stock_service.registar_movimento(
        db,
        company_id=current_user.company_id,
        produto_id=body.produto_id,
        tipo=body.tipo,
        quantidade=body.quantidade,
        armazem_destino_id=body.armazem_id,
        custo_unitario=body.custo_unitario,
        motivo=body.motivo,
        created_by=current_user.id,
    )
    await write_audit(
        db, current_user.id, current_user.company_id,
        "stock_entrada", "stock_movimento", mov.id,
        dados_novos={"produto_id": str(body.produto_id), "qtd": str(body.quantidade)},
        ip_address=req.client.host if req.client else None,
    )
    await db.commit()
    return mov


@router.post("/movimentos/saida", response_model=MovimentoResponseDTO, status_code=201)
async def movimento_saida(
    req: Request,
    body: SaidaDTO,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mov = await stock_service.registar_movimento(
        db,
        company_id=current_user.company_id,
        produto_id=body.produto_id,
        tipo=body.tipo,
        quantidade=body.quantidade,
        armazem_origem_id=body.armazem_id,
        motivo=body.motivo,
        created_by=current_user.id,
    )
    await write_audit(
        db, current_user.id, current_user.company_id,
        "stock_saida", "stock_movimento", mov.id,
        dados_novos={"produto_id": str(body.produto_id), "qtd": str(body.quantidade)},
        ip_address=req.client.host if req.client else None,
    )
    await db.commit()
    return mov


@router.post("/movimentos/transferencia", response_model=MovimentoResponseDTO, status_code=201)
async def movimento_transferencia(
    req: Request,
    body: TransferenciaDTO,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mov = await stock_service.registar_movimento(
        db,
        company_id=current_user.company_id,
        produto_id=body.produto_id,
        tipo="transferencia",
        quantidade=body.quantidade,
        armazem_origem_id=body.armazem_origem_id,
        armazem_destino_id=body.armazem_destino_id,
        motivo=body.motivo,
        created_by=current_user.id,
    )
    await write_audit(
        db, current_user.id, current_user.company_id,
        "stock_transferencia", "stock_movimento", mov.id,
        dados_novos={
            "produto_id": str(body.produto_id),
            "qtd": str(body.quantidade),
            "origem": str(body.armazem_origem_id),
            "destino": str(body.armazem_destino_id),
        },
        ip_address=req.client.host if req.client else None,
    )
    await db.commit()
    return mov


@router.post("/movimentos/{id}/estornar", response_model=MovimentoResponseDTO, status_code=201)
async def estornar_movimento(
    id: UUID,
    req: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    r = await db.execute(select(StockMovimentoModel).where(StockMovimentoModel.id == id))
    orig = r.scalar_one_or_none()
    if not orig or orig.company_id != current_user.company_id:
        raise HTTPException(404, "Movimento não encontrado")
    if orig.estornado_de:
        raise HTTPException(400, "Não é possível estornar um estorno")

    # Inverter: trocar origem ↔ destino
    inverse_kw = {
        "armazem_origem_id": orig.armazem_destino_id,
        "armazem_destino_id": orig.armazem_origem_id,
    }
    if orig.tipo in stock_service.TIPOS_ENTRADA:
        # entrada → saída ajuste (sai do destino)
        tipo_inv = "saida_ajuste"
        inverse_kw = {"armazem_origem_id": orig.armazem_destino_id,
                      "armazem_destino_id": None}
    elif orig.tipo in stock_service.TIPOS_SAIDA:
        tipo_inv = "entrada_ajuste"
        inverse_kw = {"armazem_origem_id": None,
                      "armazem_destino_id": orig.armazem_origem_id}
    else:  # transferencia
        tipo_inv = "transferencia"

    mov = await stock_service.registar_movimento(
        db,
        company_id=current_user.company_id,
        produto_id=orig.produto_id,
        tipo=tipo_inv,
        quantidade=Decimal(orig.quantidade),
        motivo=f"Estorno de {orig.id}",
        created_by=current_user.id,
        estornado_de=orig.id,
        permitir_negativo=True,
        **inverse_kw,
    )
    await write_audit(
        db, current_user.id, current_user.company_id,
        "stock_estorno", "stock_movimento", mov.id,
        dados_novos={"estornado_de": str(orig.id)},
        ip_address=req.client.host if req.client else None,
    )
    await db.commit()
    return mov


# ─── Export Excel — Kardex com cabeçalho institucional ───────────────


TIPO_LABEL = {
    "entrada_compra": "Entrada · Compra",
    "entrada_producao": "Entrada · Produção",
    "entrada_ajuste": "Entrada · Ajuste",
    "saida_venda": "Saída · Venda",
    "saida_perda": "Saída · Perda",
    "saida_ajuste": "Saída · Ajuste",
    "transferencia": "Transferência",
}


@router.get("/movimentos/export")
async def exportar_kardex_excel(
    data_de: Optional[datetime] = None,
    data_ate: Optional[datetime] = None,
    tipo: Optional[str] = None,
    produto_id: Optional[UUID] = None,
    armazem_id: Optional[UUID] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exporta o kardex (movimentos de stock) em Excel com cabeçalho
    institucional da empresa + totais por tipo, no mesmo formato de
    'Movimentos Financeiros'."""
    # Carregar movimentos com filtros
    stmt = (select(StockMovimentoModel)
            .where(StockMovimentoModel.company_id == current_user.company_id))
    if data_de: stmt = stmt.where(StockMovimentoModel.created_at >= data_de)
    if data_ate: stmt = stmt.where(StockMovimentoModel.created_at <= data_ate)
    if tipo: stmt = stmt.where(StockMovimentoModel.tipo == tipo)
    if produto_id: stmt = stmt.where(StockMovimentoModel.produto_id == produto_id)
    if armazem_id:
        stmt = stmt.where(or_(
            StockMovimentoModel.armazem_origem_id == armazem_id,
            StockMovimentoModel.armazem_destino_id == armazem_id,
        ))
    stmt = stmt.order_by(StockMovimentoModel.created_at.desc()).limit(10000)
    r = await db.execute(stmt)
    movimentos = list(r.scalars().all())

    # Lookups: produto e armazem
    pr = await db.execute(select(ProdutoModel)
                          .where(ProdutoModel.company_id == current_user.company_id))
    produtos = {str(p.id): p for p in pr.scalars().all()}
    ar = await db.execute(select(ArmazemModel)
                          .where(ArmazemModel.company_id == current_user.company_id))
    armazens = {str(a.id): a for a in ar.scalars().all()}

    # Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kardex"

    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    summary_fill = PatternFill("solid", fgColor="EFF6FF")
    summary_font = Font(bold=True, size=11)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")

    # Subtítulo período
    periodo = ""
    if data_de: periodo += f"De {data_de.strftime('%d/%m/%Y')} "
    if data_ate: periodo += f"a {data_ate.strftime('%d/%m/%Y')}"

    header_row = await aplicar_cabecalho_empresa(
        ws, db, current_user.company_id,
        titulo="Kardex · Movimentos de Stock",
        subtitulo=periodo or None,
        n_colunas=9,
    )

    headers = ["Nº", "Data", "Produto (SKU)", "Tipo", "Quantidade",
               "Origem", "Destino", "Documento", "Motivo"]
    widths = [5, 18, 32, 22, 14, 16, 16, 22, 35]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[header_row].height = 20

    totais_por_tipo: dict[str, Decimal] = {}
    total_entradas = Decimal("0")
    total_saidas = Decimal("0")

    for idx, m in enumerate(movimentos, header_row + 1):
        prod = produtos.get(str(m.produto_id))
        produto_txt = f"{prod.nome} ({prod.sku})" if prod else str(m.produto_id)[:8]
        ao = armazens.get(str(m.armazem_origem_id)) if m.armazem_origem_id else None
        ad = armazens.get(str(m.armazem_destino_id)) if m.armazem_destino_id else None
        qtd = Decimal(m.quantidade)

        row_data = [
            idx - header_row,
            m.created_at.strftime("%d/%m/%Y %H:%M") if m.created_at else "",
            produto_txt,
            TIPO_LABEL.get(m.tipo, m.tipo),
            float(qtd),
            ao.codigo if ao else "—",
            ad.codigo if ad else "—",
            f"{m.documento_ref_tipo or ''} {m.documento_ref_id or ''}".strip() or "—",
            m.motivo or "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin
            if col == 5:
                cell.number_format = "#,##0.000"
                cell.alignment = Alignment(horizontal="right")
            elif col == 4:
                if m.tipo.startswith("entrada"):
                    cell.font = Font(color="166534", bold=True)
                elif m.tipo.startswith("saida"):
                    cell.font = Font(color="991B1B", bold=True)
                else:
                    cell.font = Font(color="1E40AF", bold=True)

        totais_por_tipo[m.tipo] = totais_por_tipo.get(m.tipo, Decimal("0")) + qtd
        if m.tipo.startswith("entrada"):
            total_entradas += qtd
        elif m.tipo.startswith("saida"):
            total_saidas += qtd

    # Bloco de totais (mesmo layout do export de Movimentos Financeiros)
    blank_row = header_row + len(movimentos) + 1
    ws.row_dimensions[blank_row].height = 8
    summary_start = blank_row + 1

    rows_summary: list[tuple[str, object]] = [
        ("Total de Movimentos", len(movimentos)),
        ("Total Entradas (unidades)", float(total_entradas)),
        ("Total Saídas (unidades)", float(total_saidas)),
        ("Saldo (Entradas − Saídas)", float(total_entradas - total_saidas)),
        ("", ""),
    ]
    for tipo_k in sorted(totais_por_tipo.keys()):
        rows_summary.append(
            (TIPO_LABEL.get(tipo_k, tipo_k), float(totais_por_tipo[tipo_k]))
        )

    for i, (label, value) in enumerate(rows_summary):
        rr = summary_start + i
        lc = ws.cell(row=rr, column=7, value=label)
        lc.font = summary_font
        lc.fill = summary_fill
        lc.border = thin
        lc.alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=8)
        vc = ws.cell(row=rr, column=9, value=value)
        vc.font = summary_font
        vc.fill = summary_fill
        vc.border = thin
        if i > 0 and value != "":
            vc.number_format = "#,##0.000"
        vc.alignment = Alignment(horizontal="right")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"kardex_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


__all__ = ["router"]
