from fastapi import APIRouter, Depends, HTTPException, Request, status, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from uuid import UUID, uuid4
from typing import Optional, List
from datetime import datetime
from decimal import Decimal
import os, shutil, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.application.dtos import MovimentoCreateDTO, MovimentoUpdateDTO, MovimentoResponseDTO, PaginatedResponseDTO
from app.infrastructure.database import get_db
from app.infrastructure.database.models import MovimentoFinanceiroModel
from app.infrastructure.repositories import MovimentoRepository, FundoRepository, FornecedorRepository, ConceptoRepository, MovimentoHistoricoRepository
from app.infrastructure.auth.dependencies import get_current_user, require_financeiro, require_assistente
from app.infrastructure.audit import write_audit
from app.infrastructure.export import aplicar_cabecalho_empresa
from app.presentation.api.v1.periodos import is_periodo_fechado
from app.domain.entities import User, MovimentoFinanceiro
from app.config import settings
from sqlalchemy import select, func

from app.infrastructure.storage import get_storage_provider

router = APIRouter()

UPLOAD_DIR = settings.storage_path


@router.get("", response_model=PaginatedResponseDTO)
async def listar_movimentos(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=10, ge=1, le=100),
    fornecedor_id: Optional[UUID] = None,
    conceito_id: Optional[UUID] = None,
    tipo_movimento: Optional[str] = None,
    estado_pagamento: Optional[str] = None,
    estado_movimento: Optional[str] = None,
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    repo = MovimentoRepository(db)
    skip = (page - 1) * page_size
    total = await repo.count(
        current_user.company_id,
        fornecedor_id=fornecedor_id, conceito_id=conceito_id,
        tipo_movimento=tipo_movimento, estado_pagamento=estado_pagamento,
        estado_movimento=estado_movimento,
        data_inicio=data_inicio, data_fim=data_fim,
    )
    items = await repo.get_all(
        current_user.company_id, skip=skip, limit=page_size,
        fornecedor_id=fornecedor_id, conceito_id=conceito_id,
        tipo_movimento=tipo_movimento, estado_pagamento=estado_pagamento,
        estado_movimento=estado_movimento,
        data_inicio=data_inicio, data_fim=data_fim,
    )
    return PaginatedResponseDTO(
        items=[_to_dto(m) for m in items],
        total=total, page=page, page_size=page_size,
        total_pages=(total + page_size - 1) // page_size,
    )


# NOTE: /export MUST come before /{id} to avoid route conflict
@router.get("/export")
async def exportar_excel(
    fornecedor_id: Optional[UUID] = None,
    conceito_id: Optional[UUID] = None,
    tipo_movimento: Optional[str] = None,
    estado_pagamento: Optional[str] = None,
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mov_repo = MovimentoRepository(db)
    forn_repo = FornecedorRepository(db)
    conc_repo = ConceptoRepository(db)
    fundo_repo = FundoRepository(db)

    items = await mov_repo.get_all(
        current_user.company_id, skip=0, limit=10000,
        fornecedor_id=fornecedor_id, conceito_id=conceito_id,
        tipo_movimento=tipo_movimento, estado_pagamento=estado_pagamento,
        data_inicio=data_inicio, data_fim=data_fim,
    )
    fornecedores = await forn_repo.get_by_company(current_user.company_id)
    conceitos = await conc_repo.get_by_company(current_user.company_id)
    fundo = await fundo_repo.get_by_company(current_user.company_id)
    forn_map = {str(f.id): f.nome for f in fornecedores}
    conc_map = {str(c.id): c.nome for c in conceitos}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimentos"

    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    summary_fill = PatternFill("solid", fgColor="EFF6FF")
    summary_font = Font(bold=True, size=11)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    # Cabeçalho da empresa
    periodo_txt = ""
    if data_inicio: periodo_txt += f"De {data_inicio.strftime('%d/%m/%Y')} "
    if data_fim: periodo_txt += f"a {data_fim.strftime('%d/%m/%Y')}"
    header_row = await aplicar_cabecalho_empresa(
        ws, db, current_user.company_id,
        titulo="Movimentos Financeiros",
        subtitulo=periodo_txt or None,
        n_colunas=10,
    )

    headers = ["Nº", "Data", "Fornecedor", "Conceito", "Fatura Proforma", "Fatura Recibo", "Tipo", "Estado", "Valor (AOA)", "Observações"]
    col_widths = [5, 14, 30, 25, 18, 18, 12, 12, 20, 35]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[header_row].height = 20

    total_entradas = Decimal("0")
    total_saidas = Decimal("0")
    for idx, m in enumerate(items, header_row + 1):
        valor = Decimal(str(m.valor))
        row_data = [
            idx - header_row,
            m.data.strftime("%d/%m/%Y") if m.data else "",
            forn_map.get(str(m.fornecedor_id), ""),
            conc_map.get(str(m.conceito_id), ""),
            m.fatura_proforma or "",
            m.fatura_recibo or "",
            "Entrada" if m.tipo_movimento == "entrada" else "Saída",
            m.estado_pagamento.capitalize(),
            float(valor),
            m.observacoes or "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin
            if col == 9:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col == 7:
                cell.font = Font(color="166534", bold=True) if m.tipo_movimento == "entrada" else Font(color="991B1B", bold=True)
        if m.tipo_movimento == "entrada":
            total_entradas += valor
        else:
            total_saidas += valor

    blank_row = header_row + len(items) + 1
    ws.row_dimensions[blank_row].height = 8
    summary_start = blank_row + 1
    valor_disponivel = float(fundo.valor_disponivel) if fundo else 0.0
    acumulado = float(fundo.acumulado) if fundo else 0.0
    saldo_atual = float(fundo.saldo_atual) if fundo else 0.0

    for i, (label, value) in enumerate([
        ("Total de Registos", len(items)),
        ("Total Entradas (AOA)", float(total_entradas)),
        ("Total Saídas (AOA)", float(total_saidas)),
        ("Saldo (Entradas − Saídas) (AOA)", float(total_entradas - total_saidas)),
        ("", ""),
        ("Valor Disponível (Fundo) (AOA)", valor_disponivel),
        ("Acumulado (Saídas Pagas) (AOA)", acumulado),
        ("Saldo Actual (AOA)", saldo_atual),
    ]):
        r = summary_start + i
        lc = ws.cell(row=r, column=8, value=label)
        lc.font = summary_font; lc.fill = summary_fill; lc.border = thin
        lc.alignment = Alignment(horizontal="right")
        vc = ws.cell(row=r, column=9, value=value)
        vc.font = summary_font; vc.fill = summary_fill; vc.border = thin
        if i > 0:
            vc.number_format = '#,##0.00'
        vc.alignment = Alignment(horizontal="right")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"movimentos_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{id}", response_model=MovimentoResponseDTO)
async def obter_movimento(
    id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    repo = MovimentoRepository(db)
    m = await repo.get_by_id(id)
    if not m or m.company_id != current_user.company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Movimento não encontrado")
    return _to_dto(m)


@router.get("/{id}/historico")
async def listar_historico_movimento(
    id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    mov_repo = MovimentoRepository(db)
    existing = await mov_repo.get_by_id(id)
    if not existing or existing.company_id != current_user.company_id:
        raise HTTPException(status_code=404, detail="Movimento não encontrado")
    hist_repo = MovimentoHistoricoRepository(db)
    return await hist_repo.listar(id)


async def _verificar_fatura_duplicada(
    db: AsyncSession,
    company_id,
    fatura_proforma: Optional[str],
    fatura_recibo: Optional[str],
    excluir_id: Optional[UUID] = None,
) -> list[dict]:
    """Devolve lista de movimentos existentes com a mesma proforma/recibo."""
    from sqlalchemy import or_, and_
    proforma = (fatura_proforma or "").strip()
    recibo = (fatura_recibo or "").strip()
    if not proforma and not recibo:
        return []
    conds = []
    if proforma:
        conds.append(MovimentoFinanceiroModel.fatura_proforma == proforma)
    if recibo:
        conds.append(MovimentoFinanceiroModel.fatura_recibo == recibo)
    filters = [
        MovimentoFinanceiroModel.company_id == company_id,
        MovimentoFinanceiroModel.deleted_at == None,
        or_(*conds),
    ]
    if excluir_id:
        filters.append(MovimentoFinanceiroModel.id != excluir_id)
    r = await db.execute(select(MovimentoFinanceiroModel).where(and_(*filters)).limit(5))
    return [{
        "id": str(m.id),
        "codigo": m.codigo,
        "data": m.data.isoformat() if m.data else None,
        "valor": float(m.valor),
        "tipo_movimento": m.tipo_movimento,
        "estado_pagamento": m.estado_pagamento,
        "fatura_proforma": m.fatura_proforma,
        "fatura_recibo": m.fatura_recibo,
        "match": "proforma" if proforma and m.fatura_proforma == proforma else "recibo",
    } for m in r.scalars().all()]


@router.get("/check-fatura")
async def check_fatura(
    fatura_proforma: Optional[str] = None,
    fatura_recibo: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Verifica se já existe um movimento com a fatura informada."""
    existentes = await _verificar_fatura_duplicada(
        db, current_user.company_id, fatura_proforma, fatura_recibo,
    )
    return {"duplicado": len(existentes) > 0, "movimentos": existentes}


@router.post("", response_model=MovimentoResponseDTO, status_code=status.HTTP_201_CREATED)
async def criar_movimento(
    req: Request,
    body: MovimentoCreateDTO,
    allow_duplicate: bool = Query(default=False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_assistente),
):
    from app.infrastructure.repositories import ClienteRepository
    forn_repo = FornecedorRepository(db)
    conc_repo = ConceptoRepository(db)
    fundo_repo = FundoRepository(db)
    mov_repo = MovimentoRepository(db)
    cli_repo = ClienteRepository(db)

    # Validação consoante o tipo de movimento
    if body.tipo_movimento == "saida":
        if not body.fornecedor_id:
            raise HTTPException(400, "Fornecedor obrigatório para Saída")
        fornecedor = await forn_repo.get_by_id(body.fornecedor_id)
        if not fornecedor or fornecedor.company_id != current_user.company_id:
            raise HTTPException(400, "Fornecedor inválido")
    elif body.tipo_movimento == "entrada":
        if not body.cliente_id:
            raise HTTPException(400, "Cliente obrigatório para Entrada")
        cliente = await cli_repo.get_by_id(body.cliente_id)
        if not cliente or cliente.company_id != current_user.company_id:
            raise HTTPException(400, "Cliente inválido")

    conceito = await conc_repo.get_by_id(body.conceito_id)
    if not conceito or conceito.company_id != current_user.company_id:
        raise HTTPException(status_code=400, detail="Conceito inválido")

    if await is_periodo_fechado(db, current_user.company_id, body.data):
        raise HTTPException(status_code=403, detail=f"Período {body.data.year}-{body.data.month:02d} encontra-se fechado. Para reabrir, contacte o Administrador/a.")

    # Validar duplicação de Fatura Proforma / Recibo
    if not allow_duplicate:
        existentes = await _verificar_fatura_duplicada(
            db, current_user.company_id, body.fatura_proforma, body.fatura_recibo,
        )
        if existentes:
            raise HTTPException(status_code=409, detail={
                "code": "DUPLICATE_FATURA",
                "message": "Já existe um movimento com esta fatura.",
                "movimentos": existentes,
            })

    # Estado fatura opcional na criação. Se não vier (ou vier "criado"/""), fica vazio.
    estado_pag_raw = (body.estado_pagamento or "").strip()
    if estado_pag_raw in ("", "criado"):
        estado_pag = ""  # sem estado de fatura → estado_movimento = 'criado'
    else:
        estado_pag = estado_pag_raw
        _validar_estado_pagamento(body.tipo_movimento, estado_pag)

    if body.tipo_movimento == "saida":
        fundo = await fundo_repo.get_by_company_and_tipo(current_user.company_id, body.fundo_tipo)
        if not fundo or Decimal(str(fundo.valor_disponivel or 0)) <= 0:
            raise HTTPException(
                status_code=400,
                detail=f"O fundo {body.fundo_tipo} não tem valor disponível carregado. Carregue o fundo antes de registar uma saída.",
            )

    codigo = await _gerar_codigo(db, current_user.company_id)
    estado_mov = _calcular_estado_movimento(estado_pag, body.tipo_movimento)

    entity = MovimentoFinanceiro(
        id=uuid4(),
        company_id=current_user.company_id,
        data=body.data,
        fornecedor_id=body.fornecedor_id if body.tipo_movimento == "saida" else None,
        cliente_id=body.cliente_id if body.tipo_movimento == "entrada" else None,
        conceito_id=body.conceito_id,
        fatura_proforma=body.fatura_proforma or "",
        valor=body.valor,
        fatura_recibo=body.fatura_recibo or "",
        observacoes=body.observacoes or "",
        tipo_movimento=body.tipo_movimento,
        estado_pagamento=estado_pag,
        estado_movimento=estado_mov,
        fundo_tipo=body.fundo_tipo,
        codigo=codigo,
        created_by=current_user.id,
    )
    created = await mov_repo.create(entity)

    hist_repo = MovimentoHistoricoRepository(db)
    await hist_repo.registar(created.id, current_user.company_id, current_user.id, {
        "codigo": (None, codigo),
        "estado_pagamento": (None, estado_pag),
        "estado_movimento": (None, estado_mov),
        "valor": (None, str(body.valor)),
        "tipo_movimento": (None, body.tipo_movimento),
        "fundo_tipo": (None, body.fundo_tipo),
    })

    await _recalcular_saldos(fundo_repo, mov_repo, current_user.company_id)
    await write_audit(
        db, current_user.id, current_user.company_id,
        "criado", "movimento", created.id,
        dados_novos=_mov_dict(created),
        ip_address=req.client.host if req.client else None,
    )
    await db.commit()
    fresh = await mov_repo.get_by_id(created.id)
    return _to_dto(fresh)


@router.put("/{id}", response_model=MovimentoResponseDTO)
async def atualizar_movimento(
    id: UUID,
    req: Request,
    body: MovimentoUpdateDTO,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_financeiro),
):
    mov_repo = MovimentoRepository(db)
    fundo_repo = FundoRepository(db)
    existing = await mov_repo.get_by_id(id)
    if not existing or existing.company_id != current_user.company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Movimento não encontrado")

    data_ref = body.data or existing.data
    if data_ref and await is_periodo_fechado(db, current_user.company_id, data_ref):
        raise HTTPException(status_code=403, detail=f"Período {data_ref.year}-{data_ref.month:02d} encontra-se fechado. Para reabrir, contacte o Administrador/a.")

    dados_ant = _mov_dict(existing)

    eff_tipo = body.tipo_movimento or existing.tipo_movimento
    eff_estado_pag = body.estado_pagamento or existing.estado_pagamento
    existing_fundo_tipo = existing.fundo_tipo or "BCS"

    _validar_estado_pagamento(eff_tipo, eff_estado_pag)
    novo_estado_mov = _calcular_estado_movimento(eff_estado_pag, eff_tipo)

    diff: dict = {}
    if body.estado_pagamento is not None and body.estado_pagamento != existing.estado_pagamento:
        diff["estado_pagamento"] = (existing.estado_pagamento, body.estado_pagamento)
    cur_estado_mov = existing.estado_movimento or "criado"
    if novo_estado_mov != cur_estado_mov:
        diff["estado_movimento"] = (cur_estado_mov, novo_estado_mov)
    if body.valor is not None and str(body.valor) != str(existing.valor):
        diff["valor"] = (str(existing.valor), str(body.valor))
    if body.tipo_movimento is not None and body.tipo_movimento != existing.tipo_movimento:
        diff["tipo_movimento"] = (existing.tipo_movimento, body.tipo_movimento)
    if body.fundo_tipo is not None and body.fundo_tipo != existing_fundo_tipo:
        diff["fundo_tipo"] = (existing_fundo_tipo, body.fundo_tipo)
    if body.fornecedor_id is not None and str(body.fornecedor_id) != str(existing.fornecedor_id):
        diff["fornecedor_id"] = (str(existing.fornecedor_id), str(body.fornecedor_id))
    if body.conceito_id is not None and str(body.conceito_id) != str(existing.conceito_id):
        diff["conceito_id"] = (str(existing.conceito_id), str(body.conceito_id))
    if body.fatura_proforma is not None and body.fatura_proforma != existing.fatura_proforma:
        diff["fatura_proforma"] = (existing.fatura_proforma, body.fatura_proforma)
    if body.fatura_recibo is not None and body.fatura_recibo != existing.fatura_recibo:
        diff["fatura_recibo"] = (existing.fatura_recibo, body.fatura_recibo)
    if body.observacoes is not None and body.observacoes != existing.observacoes:
        diff["observacoes"] = (existing.observacoes, body.observacoes)

    update_entity = MovimentoFinanceiro(
        id=existing.id,
        company_id=existing.company_id,
        data=body.data,
        fornecedor_id=body.fornecedor_id,
        cliente_id=body.cliente_id,
        conceito_id=body.conceito_id,
        valor=body.valor,
        tipo_movimento=body.tipo_movimento,
        estado_pagamento=body.estado_pagamento,
        estado_movimento=novo_estado_mov,
        fundo_tipo=body.fundo_tipo,
        fatura_proforma=body.fatura_proforma,
        fatura_recibo=body.fatura_recibo,
        observacoes=body.observacoes,
        created_by=existing.created_by,
        codigo=getattr(existing, "codigo", None),
    )
    updated = await mov_repo.update(id, update_entity)

    if diff:
        hist_repo = MovimentoHistoricoRepository(db)
        await hist_repo.registar(id, current_user.company_id, current_user.id, diff)

    await _recalcular_saldos(fundo_repo, mov_repo, current_user.company_id)
    await write_audit(
        db, current_user.id, current_user.company_id,
        "atualizado", "movimento", id,
        dados_anteriores=dados_ant,
        dados_novos=_mov_dict(updated),
        ip_address=req.client.host if req.client else None,
    )
    await db.commit()
    fresh = await mov_repo.get_by_id(id)
    return _to_dto(fresh)


@router.delete("/{id}", status_code=status.HTTP_204_NO_CONTENT)
async def eliminar_movimento(
    id: UUID,
    req: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_financeiro),
):
    mov_repo = MovimentoRepository(db)
    fundo_repo = FundoRepository(db)
    existing = await mov_repo.get_by_id(id)
    if not existing or existing.company_id != current_user.company_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Movimento não encontrado")
    if existing.data and await is_periodo_fechado(db, current_user.company_id, existing.data):
        raise HTTPException(status_code=403, detail=f"Período {existing.data.year}-{existing.data.month:02d} encontra-se fechado. Para reabrir, contacte o Administrador/a.")
    dados_ant = _mov_dict(existing)
    await mov_repo.delete(id)
    await _recalcular_saldos(fundo_repo, mov_repo, current_user.company_id)
    await write_audit(
        db, current_user.id, current_user.company_id,
        "eliminado", "movimento", id,
        dados_anteriores=dados_ant,
        ip_address=req.client.host if req.client else None,
    )
    await db.commit()


@router.post("/{id}/comprovativo")
async def upload_comprovativo(
    id: UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_assistente),
):
    mov_repo = MovimentoRepository(db)
    existing = await mov_repo.get_by_id(id)
    if not existing or existing.company_id != current_user.company_id:
        raise HTTPException(status_code=404, detail="Movimento não encontrado")
    ALLOWED_TYPES = {
        "application/pdf",
        "image/jpeg", "image/jpg", "image/png",
        "image/gif", "image/webp", "image/bmp",
        "image/tiff", "image/svg+xml",
    }
    ALLOWED_EXT = {".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tiff", ".tif", ".svg"}
    MAX_SIZE = 10 * 1024 * 1024  # 10 MB

    ext = os.path.splitext(file.filename or "")[1].lower()
    content_type = (file.content_type or "").lower()
    if ext not in ALLOWED_EXT and content_type not in ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de ficheiro não permitido. São aceites PDF e imagens (JPG, PNG, GIF, WebP, BMP, TIFF, SVG).")

    content = await file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(status_code=400, detail="Ficheiro demasiado grande. Limite máximo: 10 MB.")

    filename = f"comprovativos/{id}_{file.filename}"
    storage = get_storage_provider()
    import io as _io
    await storage.upload(filename, _io.BytesIO(content), content_type=file.content_type or "application/octet-stream")

    result = await db.execute(select(MovimentoFinanceiroModel).where(MovimentoFinanceiroModel.id == id))
    model = result.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=404, detail="Movimento não encontrado")
    model.comprovativo_pagamento = filename
    model.updated_at = datetime.utcnow()
    await db.commit()
    url = storage.public_url(filename)
    return {"filename": filename, "url": url}


@router.delete("/{id}/comprovativo", status_code=status.HTTP_204_NO_CONTENT)
async def delete_comprovativo(
    id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_financeiro),
):
    result = await db.execute(select(MovimentoFinanceiroModel).where(MovimentoFinanceiroModel.id == id))
    model = result.scalar_one_or_none()
    if not model or str(model.company_id) != str(current_user.company_id):
        raise HTTPException(status_code=404, detail="Movimento não encontrado")
    if not model.comprovativo_pagamento:
        raise HTTPException(status_code=404, detail="Nenhum comprovativo anexado")
    storage = get_storage_provider()
    await storage.delete(model.comprovativo_pagamento)
    model.comprovativo_pagamento = None
    model.updated_at = datetime.utcnow()
    await db.commit()


_ESTADOS_VALIDOS = {
    "entrada": {"pendente", "pago_parcial", "pago_total"},
    "saida":   {"pendente", "pago", "cancelado", "devolvido"},
}

_ESTADOS_FECHADO = {"pago", "pago_total"}
_ESTADOS_PENDENTE = {"pendente", "pago_parcial"}


@router.post("/bulk-update")
async def bulk_update_estado(
    body: dict,
    req: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_financeiro),
):
    """Actualiza estado_pagamento de N movimentos em massa.
    Body: { "ids": ["uuid1", ...], "estado_pagamento": "pago" }
    """
    ids = body.get("ids", [])
    novo_estado = body.get("estado_pagamento")
    if not ids or not isinstance(ids, list):
        raise HTTPException(400, "Campo 'ids' obrigatório (lista de UUIDs).")
    if not novo_estado:
        raise HTTPException(400, "Campo 'estado_pagamento' obrigatório.")

    mov_repo = MovimentoRepository(db)
    fundo_repo = FundoRepository(db)
    actualizados = 0
    erros: list[dict] = []

    for raw_id in ids:
        try:
            mov_id = UUID(raw_id) if isinstance(raw_id, str) else raw_id
            existing = await mov_repo.get_by_id(mov_id)
            if not existing or existing.company_id != current_user.company_id:
                erros.append({"id": str(raw_id), "erro": "Não encontrado"})
                continue
            try:
                _validar_estado_pagamento(existing.tipo_movimento, novo_estado)
            except HTTPException as ex:
                erros.append({"id": str(raw_id), "erro": ex.detail})
                continue
            if existing.estado_pagamento == novo_estado:
                continue
            antes = existing.estado_pagamento
            existing.estado_pagamento = novo_estado
            existing.estado_movimento = _calcular_estado_movimento(novo_estado, existing.tipo_movimento)
            await mov_repo.update(mov_id, existing)
            await write_audit(
                db, current_user.id, current_user.company_id,
                "actualizado_bulk", "movimento", mov_id,
                dados_anteriores={"estado_pagamento": antes},
                dados_novos={"estado_pagamento": novo_estado},
                ip_address=req.client.host if req.client else None,
            )
            actualizados += 1
        except Exception as e:
            erros.append({"id": str(raw_id), "erro": str(e)})

    if actualizados > 0:
        await _recalcular_saldos(fundo_repo, mov_repo, current_user.company_id)
    await db.commit()
    return {"actualizados": actualizados, "total_pedidos": len(ids), "erros": erros}


@router.post("/bulk-delete")
async def bulk_delete(
    body: dict,
    req: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_financeiro),
):
    """Soft-delete de N movimentos. Body: { "ids": ["uuid1", ...] }"""
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(400, "Campo 'ids' obrigatório.")

    mov_repo = MovimentoRepository(db)
    fundo_repo = FundoRepository(db)
    eliminados = 0
    erros: list[dict] = []

    for raw_id in ids:
        try:
            mov_id = UUID(raw_id) if isinstance(raw_id, str) else raw_id
            existing = await mov_repo.get_by_id(mov_id)
            if not existing or existing.company_id != current_user.company_id:
                erros.append({"id": str(raw_id), "erro": "Não encontrado"})
                continue
            dados_ant = _mov_dict(existing)
            await mov_repo.delete(mov_id)
            await write_audit(
                db, current_user.id, current_user.company_id,
                "eliminado_bulk", "movimento", mov_id,
                dados_anteriores=dados_ant,
                ip_address=req.client.host if req.client else None,
            )
            eliminados += 1
        except Exception as e:
            erros.append({"id": str(raw_id), "erro": str(e)})

    if eliminados > 0:
        await _recalcular_saldos(fundo_repo, mov_repo, current_user.company_id)
    await db.commit()
    return {"eliminados": eliminados, "total_pedidos": len(ids), "erros": erros}


def _validar_estado_pagamento(tipo: str, estado: str):
    validos = _ESTADOS_VALIDOS.get(tipo, set())
    if estado not in validos:
        raise HTTPException(400, f"Estado '{estado}' inválido para movimento do tipo '{tipo}'. Valores aceites: {sorted(validos)}")


def _calcular_estado_movimento(estado_pagamento: str, tipo_movimento: str, tem_anexo: bool = False) -> str:
    """Workflow do movimento:
    - sem estado de fatura definido → 'criado'
    - com estado de fatura definido  → 'pendente'
    - com pelo menos um anexo → 'fechado' (sobrepõe tudo)
    """
    if tem_anexo:
        return "fechado"
    if not estado_pagamento or estado_pagamento.strip() == "":
        return "criado"
    return "pendente"


async def _gerar_codigo(db: AsyncSession, company_id) -> str:
    result = await db.execute(
        select(func.max(MovimentoFinanceiroModel.codigo))
        .where(MovimentoFinanceiroModel.company_id == company_id)
    )
    last = result.scalar_one_or_none()
    if last and last.startswith("Mov-"):
        try:
            seq = int(last[4:]) + 1
        except ValueError:
            seq = 1
    else:
        count_result = await db.execute(
            select(func.count(MovimentoFinanceiroModel.id))
            .where(MovimentoFinanceiroModel.company_id == company_id)
        )
        seq = (count_result.scalar_one() or 0) + 1
    return f"Mov-{seq:06d}"


def _diff_campos(antes, depois: dict) -> dict:
    changes = {}
    for campo, val_novo in depois.items():
        val_ant = getattr(antes, campo, None)
        if str(val_ant) != str(val_novo) and val_novo is not None:
            changes[campo] = (val_ant, val_novo)
    return changes


def _mov_dict(m) -> dict:
    return {
        "id": str(m.id),
        "data": m.data.isoformat() if m.data else None,
        "fornecedor_id": str(m.fornecedor_id) if m.fornecedor_id else None,
        "cliente_id": str(m.cliente_id) if getattr(m, "cliente_id", None) else None,
        "conceito_id": str(m.conceito_id),
        "valor": str(m.valor),
        "tipo_movimento": m.tipo_movimento,
        "estado_pagamento": m.estado_pagamento,
        "fatura_proforma": m.fatura_proforma,
        "fatura_recibo": m.fatura_recibo,
        "observacoes": m.observacoes,
    }


async def _recalcular_saldos(fundo_repo: FundoRepository, mov_repo: MovimentoRepository, company_id):
    for tipo in ("BCS", "BFA"):
        acumulado = await mov_repo.sum_by_tipo(company_id, "saida", "pago", fundo_tipo=tipo)
        total_entradas = await mov_repo.sum_by_tipo(company_id, "entrada", estado=None, fundo_tipo=tipo)
        fundo = await fundo_repo.get_by_company_and_tipo(company_id, tipo)
        if fundo:
            disponivel = Decimal(str(fundo.valor_disponivel or 0))
            saldo = disponivel + Decimal(str(total_entradas)) - Decimal(str(acumulado))
            await fundo_repo.update_saldos_by_tipo(company_id, tipo, Decimal(str(acumulado)), saldo)


def _to_dto(m: MovimentoFinanceiro) -> MovimentoResponseDTO:
    return MovimentoResponseDTO(
        id=m.id, company_id=m.company_id, data=m.data,
        fornecedor_id=m.fornecedor_id, cliente_id=m.cliente_id, conceito_id=m.conceito_id,
        fatura_proforma=m.fatura_proforma, valor=Decimal(str(m.valor)),
        fatura_recibo=m.fatura_recibo, observacoes=m.observacoes,
        tipo_movimento=m.tipo_movimento, estado_pagamento=m.estado_pagamento,
        fundo_tipo=getattr(m, "fundo_tipo", "BCS") or "BCS",
        codigo=getattr(m, "codigo", None),
        estado_movimento=getattr(m, "estado_movimento", "criado") or "criado",
        created_by=m.created_by, created_at=m.created_at, updated_at=m.updated_at,
        deleted_at=m.deleted_at, comprovativo_pagamento=m.comprovativo_pagamento,
    )


__all__ = ["router"]
