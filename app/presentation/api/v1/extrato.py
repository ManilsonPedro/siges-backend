"""Extratos por Fornecedor / Conceito + opção "Todos" agrupado.
Exportação Excel + PDF (com watermark do logo)."""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from datetime import datetime
from decimal import Decimal
from uuid import UUID
from typing import Optional, Literal
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from pathlib import Path

from app.config import settings
from app.infrastructure.database import get_db
from app.infrastructure.database.models import (
    MovimentoFinanceiroModel, FornecedorModel, ConceptoModel,
    CompanySettingsModel,
)
from app.infrastructure.auth.dependencies import get_current_user
from app.infrastructure.export import aplicar_cabecalho_empresa
from app.infrastructure.export.pdf import render_extrato_pdf
from app.domain.entities import User

router = APIRouter()


async def _carregar_movimentos(
    db: AsyncSession,
    company_id,
    fornecedor_id: Optional[UUID],
    conceito_id: Optional[UUID],
    data_inicio: Optional[datetime],
    data_fim: Optional[datetime],
) -> list[MovimentoFinanceiroModel]:
    filters = [
        MovimentoFinanceiroModel.company_id == company_id,
        MovimentoFinanceiroModel.deleted_at == None,
    ]
    if fornecedor_id:
        filters.append(MovimentoFinanceiroModel.fornecedor_id == fornecedor_id)
    if conceito_id:
        filters.append(MovimentoFinanceiroModel.conceito_id == conceito_id)
    if data_inicio:
        filters.append(MovimentoFinanceiroModel.data >= data_inicio)
    if data_fim:
        filters.append(MovimentoFinanceiroModel.data <= data_fim)
    r = await db.execute(
        select(MovimentoFinanceiroModel).where(and_(*filters))
        .order_by(MovimentoFinanceiroModel.data.desc())
    )
    return list(r.scalars().all())


def _mov_to_dict(m: MovimentoFinanceiroModel) -> dict:
    return {
        "id": str(m.id),
        "codigo": m.codigo,
        "data": m.data.isoformat() if m.data else None,
        "valor": float(m.valor),
        "tipo_movimento": m.tipo_movimento,
        "estado_pagamento": m.estado_pagamento,
        "estado_movimento": m.estado_movimento,
        "fundo_tipo": m.fundo_tipo,
        "fornecedor_id": str(m.fornecedor_id),
        "conceito_id": str(m.conceito_id),
        "fatura_proforma": m.fatura_proforma,
        "fatura_recibo": m.fatura_recibo,
        "observacoes": m.observacoes,
    }


def _calcular_totais(movimentos: list[dict]) -> dict:
    entradas = sum(m["valor"] for m in movimentos if m["tipo_movimento"] == "entrada")
    saidas = sum(m["valor"] for m in movimentos if m["tipo_movimento"] == "saida")
    return {
        "entradas": entradas,
        "saidas": saidas,
        "saldo": entradas - saidas,
        "count": len(movimentos),
    }


async def _get_company_settings(db: AsyncSession, company_id) -> dict:
    r = await db.execute(select(CompanySettingsModel).where(CompanySettingsModel.company_id == company_id))
    cs = r.scalar_one_or_none()
    if not cs:
        return {"nome": "SIGES BI JENNOS", "logo_path": None}
    return {
        "nome": cs.nome or "SIGES BI JENNOS",
        "nif": cs.nif, "morada": cs.morada, "telefone": cs.telefone,
        "email": cs.email, "logo_path": cs.logo_path,
    }


# ─────────────────────────────────────────────────────────────────────
# Endpoints — singular (1 fornecedor / 1 conceito)
# ─────────────────────────────────────────────────────────────────────

@router.get("/fornecedor/{fornecedor_id}")
async def extrato_fornecedor(
    fornecedor_id: UUID,
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rf = await db.execute(select(FornecedorModel).where(FornecedorModel.id == fornecedor_id))
    f = rf.scalar_one_or_none()
    if not f or f.company_id != current_user.company_id:
        raise HTTPException(404, "Fornecedor não encontrado")

    movimentos = await _carregar_movimentos(
        db, current_user.company_id, fornecedor_id, None, data_inicio, data_fim,
    )
    conc_ids = list({m.conceito_id for m in movimentos})
    conc_map = {}
    if conc_ids:
        rc = await db.execute(select(ConceptoModel).where(ConceptoModel.id.in_(conc_ids)))
        conc_map = {str(c.id): c.nome for c in rc.scalars().all()}

    movs_dict = [{**_mov_to_dict(m), "conceito_nome": conc_map.get(str(m.conceito_id))} for m in movimentos]

    return {
        "fornecedor": {
            "id": str(f.id), "nome": f.nome, "nif": f.nif,
            "telefone": f.telefone, "email": f.email,
            "endereco": f.endereco, "estado": f.estado,
        },
        "periodo": {
            "data_inicio": data_inicio.isoformat() if data_inicio else None,
            "data_fim": data_fim.isoformat() if data_fim else None,
        },
        "movimentos": movs_dict,
        "totais": _calcular_totais(movs_dict),
    }


@router.get("/conceito/{conceito_id}")
async def extrato_conceito(
    conceito_id: UUID,
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rc = await db.execute(select(ConceptoModel).where(ConceptoModel.id == conceito_id))
    c = rc.scalar_one_or_none()
    if not c or c.company_id != current_user.company_id:
        raise HTTPException(404, "Conceito não encontrado")

    movimentos = await _carregar_movimentos(
        db, current_user.company_id, None, conceito_id, data_inicio, data_fim,
    )
    forn_ids = list({m.fornecedor_id for m in movimentos})
    forn_map = {}
    if forn_ids:
        rf = await db.execute(select(FornecedorModel).where(FornecedorModel.id.in_(forn_ids)))
        forn_map = {str(f.id): f.nome for f in rf.scalars().all()}

    movs_dict = [{**_mov_to_dict(m), "fornecedor_nome": forn_map.get(str(m.fornecedor_id))} for m in movimentos]

    return {
        "conceito": {
            "id": str(c.id), "nome": c.nome,
            "descricao": c.descricao, "estado": c.estado,
        },
        "periodo": {
            "data_inicio": data_inicio.isoformat() if data_inicio else None,
            "data_fim": data_fim.isoformat() if data_fim else None,
        },
        "movimentos": movs_dict,
        "totais": _calcular_totais(movs_dict),
    }


# ─────────────────────────────────────────────────────────────────────
# Endpoints — TODOS (agrupado alfabeticamente)
# ─────────────────────────────────────────────────────────────────────

@router.get("/fornecedores/todos")
async def extrato_todos_fornecedores(
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Extrato com TODOS os fornecedores agrupados (ordem alfabética)."""
    rf = await db.execute(
        select(FornecedorModel)
        .where(and_(FornecedorModel.company_id == current_user.company_id, FornecedorModel.deleted_at == None))
        .order_by(FornecedorModel.nome.asc())
    )
    fornecedores = list(rf.scalars().all())

    movimentos = await _carregar_movimentos(
        db, current_user.company_id, None, None, data_inicio, data_fim,
    )
    conc_ids = list({m.conceito_id for m in movimentos})
    conc_map = {}
    if conc_ids:
        rc = await db.execute(select(ConceptoModel).where(ConceptoModel.id.in_(conc_ids)))
        conc_map = {str(c.id): c.nome for c in rc.scalars().all()}

    movs_by_forn: dict[str, list[dict]] = {}
    for m in movimentos:
        key = str(m.fornecedor_id)
        movs_by_forn.setdefault(key, []).append({**_mov_to_dict(m), "conceito_nome": conc_map.get(str(m.conceito_id))})

    grupos = []
    total_entradas = 0.0
    total_saidas = 0.0
    total_count = 0
    for f in fornecedores:
        movs = movs_by_forn.get(str(f.id), [])
        if not movs:
            continue
        t = _calcular_totais(movs)
        grupos.append({
            "label": f.nome,
            "entidade": {"id": str(f.id), "nome": f.nome, "nif": f.nif},
            "movimentos": movs,
            "totais": t,
        })
        total_entradas += t["entradas"]
        total_saidas += t["saidas"]
        total_count += t["count"]

    return {
        "periodo": {
            "data_inicio": data_inicio.isoformat() if data_inicio else None,
            "data_fim": data_fim.isoformat() if data_fim else None,
        },
        "grupos": grupos,
        "totais_gerais": {
            "entradas": total_entradas,
            "saidas": total_saidas,
            "saldo": total_entradas - total_saidas,
            "count": total_count,
            "n_grupos": len(grupos),
        },
    }


@router.get("/conceitos/todos")
async def extrato_todos_conceitos(
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Extrato com TODOS os conceitos agrupados (ordem alfabética)."""
    rc = await db.execute(
        select(ConceptoModel)
        .where(and_(ConceptoModel.company_id == current_user.company_id, ConceptoModel.deleted_at == None))
        .order_by(ConceptoModel.nome.asc())
    )
    conceitos = list(rc.scalars().all())

    movimentos = await _carregar_movimentos(
        db, current_user.company_id, None, None, data_inicio, data_fim,
    )
    forn_ids = list({m.fornecedor_id for m in movimentos})
    forn_map = {}
    if forn_ids:
        rf = await db.execute(select(FornecedorModel).where(FornecedorModel.id.in_(forn_ids)))
        forn_map = {str(f.id): f.nome for f in rf.scalars().all()}

    movs_by_conc: dict[str, list[dict]] = {}
    for m in movimentos:
        key = str(m.conceito_id)
        movs_by_conc.setdefault(key, []).append({**_mov_to_dict(m), "fornecedor_nome": forn_map.get(str(m.fornecedor_id))})

    grupos = []
    total_entradas = 0.0
    total_saidas = 0.0
    total_count = 0
    for c in conceitos:
        movs = movs_by_conc.get(str(c.id), [])
        if not movs:
            continue
        t = _calcular_totais(movs)
        grupos.append({
            "label": c.nome,
            "entidade": {"id": str(c.id), "nome": c.nome},
            "movimentos": movs,
            "totais": t,
        })
        total_entradas += t["entradas"]
        total_saidas += t["saidas"]
        total_count += t["count"]

    return {
        "periodo": {
            "data_inicio": data_inicio.isoformat() if data_inicio else None,
            "data_fim": data_fim.isoformat() if data_fim else None,
        },
        "grupos": grupos,
        "totais_gerais": {
            "entradas": total_entradas,
            "saidas": total_saidas,
            "saldo": total_entradas - total_saidas,
            "count": total_count,
            "n_grupos": len(grupos),
        },
    }


# ─────────────────────────────────────────────────────────────────────
# Export — Excel + PDF
# ─────────────────────────────────────────────────────────────────────

async def _gerar_excel(
    db: AsyncSession,
    company_id,
    *,
    titulo: str,
    subtitulo: Optional[str],
    entidade_rows: list[tuple[str, str]],
    grupos: list[dict],  # [{label, movimentos, totais}, ...]
    totais_gerais: dict,
    tipo_aux: str,  # "Conceito" ou "Fornecedor"
    chave_aux: str,  # "conceito_nome" ou "fornecedor_nome"
    mostrar_entradas: bool = True,
    mostrar_saidas: bool = True,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato"

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="0B3B6F")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    row = await aplicar_cabecalho_empresa(ws, db, company_id, titulo=titulo, subtitulo=subtitulo, n_colunas=7)

    # Detalhes da entidade (se único)
    for label, value in entidade_rows:
        if not value: continue
        c1 = ws.cell(row=row, column=2, value=f"{label}:")
        c1.font = Font(bold=True, color="0B3B6F", size=10)
        c2 = ws.cell(row=row, column=3, value=value)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        row += 1
    if entidade_rows:
        row += 1

    # Cabeçalho tabela
    headers = ["#", "Código", "Data", tipo_aux, "Tipo", "Fundo", "Valor (AOA)"]
    widths = [4, 12, 12, 30, 10, 8, 16]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = thin
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 18
    row += 1

    is_agrupado = len(grupos) > 1

    for g in grupos:
        if is_agrupado:
            grp_cell = ws.cell(row=row, column=1, value=f"▸ {g['label']}  ·  {g['totais']['count']} mov.")
            grp_cell.font = Font(bold=True, color="0B3B6F", size=11)
            grp_cell.fill = PatternFill("solid", fgColor="EAF1F9")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            row += 1

        for idx, m in enumerate(g["movimentos"], 1):
            tipo_color = "166534" if m["tipo_movimento"] == "entrada" else "991B1B"
            data_lines = [
                idx,
                m.get("codigo") or "—",
                (m.get("data") or "")[:10],
                m.get(chave_aux) or "—",
                "Entrada" if m["tipo_movimento"] == "entrada" else "Saída",
                m.get("fundo_tipo") or "BCS",
                float(m["valor"]),
            ]
            for col, val in enumerate(data_lines, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin
                cell.font = Font(size=9)
                if col == 5:
                    cell.font = Font(bold=True, color=tipo_color, size=9)
                if col == 6:
                    cell.alignment = center
                if col == 7:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
            row += 1

        if is_agrupado:
            t = g["totais"]
            sub_fill = PatternFill("solid", fgColor="F9FAFB")
            for col in range(1, 8):
                c = ws.cell(row=row, column=col)
                c.fill = sub_fill; c.border = thin
            ws.cell(row=row, column=6, value=f"Subtotal {g['label']}:").font = Font(bold=True, size=9)
            ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
            val_cell = ws.cell(row=row, column=7, value=t["saldo"])
            val_cell.font = Font(bold=True, size=9)
            val_cell.number_format = '#,##0.00'
            val_cell.alignment = Alignment(horizontal="right")
            row += 1

    # Total geral
    row += 1
    total_fill = PatternFill("solid", fgColor="0B3B6F")
    total_font = Font(bold=True, color="FFFFFF", size=11)
    totais_rows: list[tuple[str, float | int]] = []
    if mostrar_entradas:
        totais_rows.append(("Total Entradas (AOA)", totais_gerais["entradas"]))
    if mostrar_saidas:
        totais_rows.append(("Total Saídas (AOA)", totais_gerais["saidas"]))
    if mostrar_entradas and mostrar_saidas:
        totais_rows.append(("SALDO (AOA)", totais_gerais["saldo"]))
    totais_rows.append(("Total Movimentos", totais_gerais["count"]))
    for label, value in totais_rows:
        lc = ws.cell(row=row, column=6, value=label)
        lc.font = total_font; lc.fill = total_fill; lc.border = thin
        lc.alignment = Alignment(horizontal="right")
        vc = ws.cell(row=row, column=7, value=value)
        vc.font = total_font; vc.fill = total_fill; vc.border = thin
        if isinstance(value, float):
            vc.number_format = '#,##0.00'
        vc.alignment = Alignment(horizontal="right")
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_pdf(
    company_settings: dict,
    *,
    titulo: str,
    subtitulo: Optional[str],
    entidade: Optional[dict],
    tipo_aux: str,
    grupos: list[dict],
    totais_gerais: dict,
    chave_aux: str,
    mostrar_entradas: bool = True,
    mostrar_saidas: bool = True,
) -> bytes:
    # Normalizar movimentos para o template PDF (esperando "aux_nome")
    grupos_pdf = []
    for g in grupos:
        movs = []
        for m in g["movimentos"]:
            mm = dict(m)
            mm["aux_nome"] = m.get(chave_aux)
            movs.append(mm)
        grupos_pdf.append({"label": g["label"], "movimentos": movs, "totais": g["totais"]})

    return render_extrato_pdf(
        titulo=titulo,
        subtitulo=subtitulo,
        empresa=company_settings,
        entidade=entidade,
        tipo_coluna_aux=tipo_aux,
        grupos=grupos_pdf,
        totais_gerais=totais_gerais,
        data_emissao=datetime.utcnow().strftime("%d/%m/%Y %H:%M"),
        mostrar_entradas=mostrar_entradas,
        mostrar_saidas=mostrar_saidas,
    )


def _periodo_subtitle(data_inicio, data_fim) -> Optional[str]:
    parts = []
    if data_inicio: parts.append(f"De {data_inicio.strftime('%d/%m/%Y')}")
    if data_fim: parts.append(f"a {data_fim.strftime('%d/%m/%Y')}")
    return " ".join(parts) if parts else None


def _stream(content: bytes, fname: str, formato: str) -> StreamingResponse:
    media = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if formato == "excel"
        else "application/pdf"
    )
    return StreamingResponse(io.BytesIO(content), media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# Export — 1 fornecedor

@router.get("/fornecedor/{fornecedor_id}/export")
async def export_extrato_fornecedor(
    fornecedor_id: UUID,
    formato: Literal["excel", "pdf"] = Query(default="excel"),
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    movimentos_ids: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rf = await db.execute(select(FornecedorModel).where(FornecedorModel.id == fornecedor_id))
    f = rf.scalar_one_or_none()
    if not f or f.company_id != current_user.company_id:
        raise HTTPException(404, "Fornecedor não encontrado")

    movimentos = await _carregar_movimentos(
        db, current_user.company_id, fornecedor_id, None, data_inicio, data_fim,
    )
    if movimentos_ids:
        ids_set = {x.strip() for x in movimentos_ids.split(",") if x.strip()}
        movimentos = [m for m in movimentos if str(m.id) in ids_set]

    conc_ids = list({m.conceito_id for m in movimentos})
    conc_map = {}
    if conc_ids:
        rc = await db.execute(select(ConceptoModel).where(ConceptoModel.id.in_(conc_ids)))
        conc_map = {str(c.id): c.nome for c in rc.scalars().all()}

    movs = [{**_mov_to_dict(m), "conceito_nome": conc_map.get(str(m.conceito_id))} for m in movimentos]
    totais = _calcular_totais(movs)
    grupos = [{"label": f.nome, "movimentos": movs, "totais": totais}]
    entidade = {"Nome": f.nome, "NIF": f.nif, "Telefone": f.telefone or "", "Email": f.email or "",
                "Morada": f.endereco or "", "Estado": f.estado}
    titulo = f"Extrato por Fornecedor: {f.nome}"
    subtitulo = _periodo_subtitle(data_inicio, data_fim)
    cs = await _get_company_settings(db, current_user.company_id)
    fname_base = f"extrato_fornecedor_{f.nome.replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d')}"

    if formato == "excel":
        xlsx = await _gerar_excel(
            db, current_user.company_id,
            titulo=titulo, subtitulo=subtitulo,
            entidade_rows=list(entidade.items()),
            grupos=grupos, totais_gerais=totais,
            tipo_aux="Conceito", chave_aux="conceito_nome",
        )
        return _stream(xlsx, f"{fname_base}.xlsx", "excel")
    else:
        pdf = _build_pdf(cs, titulo=titulo, subtitulo=subtitulo,
            entidade=entidade, tipo_aux="Conceito",
            grupos=grupos, totais_gerais=totais, chave_aux="conceito_nome")
        return _stream(pdf, f"{fname_base}.pdf", "pdf")


# Export — 1 conceito

@router.get("/conceito/{conceito_id}/export")
async def export_extrato_conceito(
    conceito_id: UUID,
    formato: Literal["excel", "pdf"] = Query(default="excel"),
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    movimentos_ids: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rc = await db.execute(select(ConceptoModel).where(ConceptoModel.id == conceito_id))
    c = rc.scalar_one_or_none()
    if not c or c.company_id != current_user.company_id:
        raise HTTPException(404, "Conceito não encontrado")

    movimentos = await _carregar_movimentos(
        db, current_user.company_id, None, conceito_id, data_inicio, data_fim,
    )
    if movimentos_ids:
        ids_set = {x.strip() for x in movimentos_ids.split(",") if x.strip()}
        movimentos = [m for m in movimentos if str(m.id) in ids_set]

    forn_ids = list({m.fornecedor_id for m in movimentos})
    forn_map = {}
    if forn_ids:
        rf = await db.execute(select(FornecedorModel).where(FornecedorModel.id.in_(forn_ids)))
        forn_map = {str(f.id): f.nome for f in rf.scalars().all()}

    movs = [{**_mov_to_dict(m), "fornecedor_nome": forn_map.get(str(m.fornecedor_id))} for m in movimentos]
    totais = _calcular_totais(movs)
    grupos = [{"label": c.nome, "movimentos": movs, "totais": totais}]
    entidade = {"Nome": c.nome, "Descrição": c.descricao or "", "Estado": c.estado}
    titulo = f"Extrato por Conceito: {c.nome}"
    subtitulo = _periodo_subtitle(data_inicio, data_fim)
    cs = await _get_company_settings(db, current_user.company_id)
    fname_base = f"extrato_conceito_{c.nome.replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d')}"

    if formato == "excel":
        xlsx = await _gerar_excel(
            db, current_user.company_id,
            titulo=titulo, subtitulo=subtitulo,
            entidade_rows=list(entidade.items()),
            grupos=grupos, totais_gerais=totais,
            tipo_aux="Fornecedor", chave_aux="fornecedor_nome",
        )
        return _stream(xlsx, f"{fname_base}.xlsx", "excel")
    else:
        pdf = _build_pdf(cs, titulo=titulo, subtitulo=subtitulo,
            entidade=entidade, tipo_aux="Fornecedor",
            grupos=grupos, totais_gerais=totais, chave_aux="fornecedor_nome")
        return _stream(pdf, f"{fname_base}.pdf", "pdf")


# Export — todos agrupados

@router.get("/fornecedores/todos/export")
async def export_todos_fornecedores(
    formato: Literal["excel", "pdf"] = Query(default="excel"),
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = await extrato_todos_fornecedores(data_inicio, data_fim, db, current_user)
    grupos = data["grupos"]
    totais = data["totais_gerais"]
    titulo = "Extrato — Todos os Fornecedores"
    subtitulo = _periodo_subtitle(data_inicio, data_fim)
    cs = await _get_company_settings(db, current_user.company_id)
    fname_base = f"extrato_todos_fornecedores_{datetime.utcnow().strftime('%Y%m%d')}"

    if formato == "excel":
        xlsx = await _gerar_excel(
            db, current_user.company_id,
            titulo=titulo, subtitulo=subtitulo, entidade_rows=[],
            grupos=grupos, totais_gerais=totais,
            tipo_aux="Conceito", chave_aux="conceito_nome",
        )
        return _stream(xlsx, f"{fname_base}.xlsx", "excel")
    else:
        pdf = _build_pdf(cs, titulo=titulo, subtitulo=subtitulo, entidade=None,
            tipo_aux="Conceito", grupos=grupos, totais_gerais=totais,
            chave_aux="conceito_nome")
        return _stream(pdf, f"{fname_base}.pdf", "pdf")


@router.get("/conceitos/todos/export")
async def export_todos_conceitos(
    formato: Literal["excel", "pdf"] = Query(default="excel"),
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = await extrato_todos_conceitos(data_inicio, data_fim, db, current_user)
    grupos = data["grupos"]
    totais = data["totais_gerais"]
    titulo = "Extrato — Todos os Conceitos"
    subtitulo = _periodo_subtitle(data_inicio, data_fim)
    cs = await _get_company_settings(db, current_user.company_id)
    fname_base = f"extrato_todos_conceitos_{datetime.utcnow().strftime('%Y%m%d')}"

    if formato == "excel":
        xlsx = await _gerar_excel(
            db, current_user.company_id,
            titulo=titulo, subtitulo=subtitulo, entidade_rows=[],
            grupos=grupos, totais_gerais=totais,
            tipo_aux="Fornecedor", chave_aux="fornecedor_nome",
        )
        return _stream(xlsx, f"{fname_base}.xlsx", "excel")
    else:
        pdf = _build_pdf(cs, titulo=titulo, subtitulo=subtitulo, entidade=None,
            tipo_aux="Fornecedor", grupos=grupos, totais_gerais=totais,
            chave_aux="fornecedor_nome")
        return _stream(pdf, f"{fname_base}.pdf", "pdf")


# ─────────────────────────────────────────────────────────────────────
# Extrato GERAL de Movimentos (sem filtro de entidade)
# ─────────────────────────────────────────────────────────────────────

@router.get("/movimentos")
async def extrato_movimentos(
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    fornecedor_id: Optional[UUID] = None,
    conceito_id: Optional[UUID] = None,
    tipo_movimento: Optional[Literal["entrada", "saida"]] = None,
    estado_pagamento: Optional[str] = None,
    fundo_tipo: Optional[Literal["BCS", "BFA"]] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista plana de movimentos com filtros, totais e contadores."""
    cid = current_user.company_id
    filters = [
        MovimentoFinanceiroModel.company_id == cid,
        MovimentoFinanceiroModel.deleted_at == None,
    ]
    if data_inicio: filters.append(MovimentoFinanceiroModel.data >= data_inicio)
    if data_fim: filters.append(MovimentoFinanceiroModel.data <= data_fim)
    if fornecedor_id: filters.append(MovimentoFinanceiroModel.fornecedor_id == fornecedor_id)
    if conceito_id: filters.append(MovimentoFinanceiroModel.conceito_id == conceito_id)
    if tipo_movimento: filters.append(MovimentoFinanceiroModel.tipo_movimento == tipo_movimento)
    if estado_pagamento: filters.append(MovimentoFinanceiroModel.estado_pagamento == estado_pagamento)
    if fundo_tipo: filters.append(MovimentoFinanceiroModel.fundo_tipo == fundo_tipo)

    r = await db.execute(
        select(MovimentoFinanceiroModel).where(and_(*filters))
        .order_by(MovimentoFinanceiroModel.data.desc())
    )
    movimentos = list(r.scalars().all())

    # Carregar nomes
    forn_ids = list({m.fornecedor_id for m in movimentos})
    forn_map = {}
    if forn_ids:
        rf = await db.execute(select(FornecedorModel).where(FornecedorModel.id.in_(forn_ids)))
        forn_map = {str(f.id): f.nome for f in rf.scalars().all()}

    conc_ids = list({m.conceito_id for m in movimentos})
    conc_map = {}
    if conc_ids:
        rc = await db.execute(select(ConceptoModel).where(ConceptoModel.id.in_(conc_ids)))
        conc_map = {str(c.id): c.nome for c in rc.scalars().all()}

    movs = [{
        **_mov_to_dict(m),
        "fornecedor_nome": forn_map.get(str(m.fornecedor_id)),
        "conceito_nome": conc_map.get(str(m.conceito_id)),
    } for m in movimentos]
    totais = _calcular_totais(movs)

    return {
        "periodo": {
            "data_inicio": data_inicio.isoformat() if data_inicio else None,
            "data_fim": data_fim.isoformat() if data_fim else None,
        },
        "movimentos": movs,
        "totais": totais,
    }


@router.get("/movimentos/export")
async def export_extrato_movimentos(
    formato: Literal["excel", "pdf"] = Query(default="excel"),
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    fornecedor_id: Optional[UUID] = None,
    conceito_id: Optional[UUID] = None,
    tipo_movimento: Optional[Literal["entrada", "saida"]] = None,
    estado_pagamento: Optional[str] = None,
    fundo_tipo: Optional[Literal["BCS", "BFA"]] = None,
    movimentos_ids: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = await extrato_movimentos(
        data_inicio, data_fim, fornecedor_id, conceito_id,
        tipo_movimento, estado_pagamento, fundo_tipo, db, current_user,
    )
    movs = data["movimentos"]
    if movimentos_ids:
        ids_set = {x.strip() for x in movimentos_ids.split(",") if x.strip()}
        movs = [m for m in movs if m["id"] in ids_set]

    totais = _calcular_totais(movs)
    grupos = [{"label": "Movimentos", "movimentos": movs, "totais": totais}]
    titulo = "Extrato de Movimentos"
    if tipo_movimento == "entrada": titulo = "Extrato de Entradas"
    elif tipo_movimento == "saida": titulo = "Extrato de Saídas"
    subtitulo = _periodo_subtitle(data_inicio, data_fim)
    cs = await _get_company_settings(db, current_user.company_id)
    fname_base = f"extrato_movimentos_{datetime.utcnow().strftime('%Y%m%d')}"

    # Esconder totais irrelevantes consoante o filtro de tipo
    mostrar_entradas = tipo_movimento != "saida"
    mostrar_saidas = tipo_movimento != "entrada"

    if formato == "excel":
        xlsx = await _gerar_excel(
            db, current_user.company_id,
            titulo=titulo, subtitulo=subtitulo, entidade_rows=[],
            grupos=grupos, totais_gerais=totais,
            tipo_aux="Fornecedor / Conceito", chave_aux="fornecedor_nome",
            mostrar_entradas=mostrar_entradas, mostrar_saidas=mostrar_saidas,
        )
        return _stream(xlsx, f"{fname_base}.xlsx", "excel")
    else:
        # Para PDF combinamos fornecedor + conceito numa linha
        for m in movs:
            m["aux_nome"] = f"{m.get('fornecedor_nome') or '—'} / {m.get('conceito_nome') or '—'}"
        grupos_pdf = [{"label": "Movimentos", "movimentos": movs, "totais": totais}]
        pdf = render_extrato_pdf(
            titulo=titulo, subtitulo=subtitulo,
            empresa=cs, entidade=None,
            tipo_coluna_aux="Fornecedor / Conceito",
            grupos=grupos_pdf, totais_gerais=totais,
            data_emissao=datetime.utcnow().strftime("%d/%m/%Y %H:%M"),
            mostrar_entradas=mostrar_entradas, mostrar_saidas=mostrar_saidas,
        )
        return _stream(pdf, f"{fname_base}.pdf", "pdf")


__all__ = ["router"]
