"""Helpers para gerar exports Excel/PDF com identidade da empresa."""
from pathlib import Path
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage

from app.config import settings
from app.infrastructure.database.models import CompanySettingsModel


async def aplicar_cabecalho_empresa(
    ws: Worksheet,
    db: AsyncSession,
    company_id,
    titulo: str,
    subtitulo: Optional[str] = None,
    n_colunas: int = 7,
) -> int:
    """Adiciona logo (se existir) + nome da empresa + título do relatório
    no topo da worksheet. Devolve a linha onde os dados devem começar.
    """
    r = await db.execute(select(CompanySettingsModel).where(
        CompanySettingsModel.company_id == company_id
    ))
    cs = r.scalar_one_or_none()

    # Linha 1: Nome da empresa
    nome = (cs.nome if cs and cs.nome else "Financ-BI Jennos")
    cell = ws.cell(row=1, column=2, value=nome)
    cell.font = Font(bold=True, size=16, color="0B3B6F")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n_colunas)

    # Linha 2: NIF + Morada
    info_partes = []
    if cs:
        if cs.nif: info_partes.append(f"NIF: {cs.nif}")
        if cs.morada: info_partes.append(cs.morada)
        if cs.telefone: info_partes.append(f"Tel: {cs.telefone}")
        if cs.email: info_partes.append(cs.email)
    if info_partes:
        cell = ws.cell(row=2, column=2, value=" · ".join(info_partes))
        cell.font = Font(size=9, color="4A5568")
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=n_colunas)

    # Linha 3: Título do relatório
    cell = ws.cell(row=3, column=2, value=titulo)
    cell.font = Font(bold=True, size=12, color="1E5A9C")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=n_colunas)

    # Linha 4: Subtítulo (opcional)
    if subtitulo:
        cell = ws.cell(row=4, column=2, value=subtitulo)
        cell.font = Font(size=10, italic=True, color="6B7280")
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=n_colunas)

    # Logo (se existir e for raster)
    if cs and cs.logo_path:
        logo_full = Path(settings.storage_path) / cs.logo_path
        # openpyxl não suporta SVG
        if logo_full.exists() and logo_full.suffix.lower() in {".png", ".jpg", ".jpeg"}:
            try:
                img = XLImage(str(logo_full))
                # Redimensionar
                img.width = 64
                img.height = 64
                ws.row_dimensions[1].height = 24
                ws.row_dimensions[2].height = 16
                ws.row_dimensions[3].height = 20
                ws.add_image(img, "A1")
                ws.column_dimensions["A"].width = 10
            except Exception:
                pass

    # Linha em branco antes dos dados
    start_row = 5 if subtitulo else 4
    return start_row + 1
